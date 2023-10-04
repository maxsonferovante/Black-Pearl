import datetime
from io import BytesIO
import openpyxl
from _decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy, reverse

from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView, FormView
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.views.generic import ListView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator


from .forms import CartaoConvenioVolusForm, FaturaCartaoForm, ContratoPlanoOdontologicoForm, ContratoPlanoSaudeForm, ContratoPlanoSaudeDependenteForm
from .models import CartaoConvenioVolus, FaturaCartao, ContratoPlanoOdontologico, TaxasAdministrativa,PlanoOdontologico, ContratoPlanoSaude, ValoresPorFaixa, PlanoSaude, ContratoPlanoSaudeDependente
from ..associados.models import Associado, Dependente


@method_decorator(login_required, name='dispatch')
class HomeTemplateView(TemplateView):
    template_name = 'convenios/home.html'

    def get_context_data(self, **kwargs):
        context = super(HomeTemplateView, self).get_context_data(**kwargs)
        return context


@method_decorator(login_required, name='dispatch')
class CartaoListView(ListView):
    template_name = 'convenios/listagemcartoes.html'
    model = CartaoConvenioVolus
    paginate_by = 5
    context_object_name = 'list_objs'

    def get_queryset(self):
        queryset = super().get_queryset()
        nome_pesquisado = self.request.GET.get('obj')
        if nome_pesquisado:
            queryset = queryset.filter(titular__nomecompleto__icontains=nome_pesquisado).order_by('titular')
        else:
            queryset = queryset.all().order_by('titular')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


@method_decorator(login_required, name='dispatch')
class CartaoCreateView(CreateView):
    model = CartaoConvenioVolus
    form_class = CartaoConvenioVolusForm
    template_name_suffix = "_criar_form"
    success_url = reverse_lazy('listagemcartoes')


@method_decorator(login_required, name='dispatch')
class CartaoUpdateView(UpdateView):
    model = CartaoConvenioVolus
    form_class = CartaoConvenioVolusForm
    template_name_suffix = "_criar_form"
    success_url = reverse_lazy('listagemcartoes')


@method_decorator(login_required, name='dispatch')
class CartaoDetailView(DetailView):
    template_name = 'convenios/cartao_detalhes.html'
    model = CartaoConvenioVolus

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cartao_pk = self.kwargs['pk']
        cartao = CartaoConvenioVolus.objects.get(pk=cartao_pk)
        context['cartao'] = cartao
        return context


@method_decorator(login_required, name='dispatch')
class CartaoDeleteView(DeleteView):
    model = CartaoConvenioVolus
    success_url = reverse_lazy('listagemcartoes')


@method_decorator(login_required, name='dispatch')
class FaturaCreateView(CreateView):
    model = FaturaCartao
    form_class = FaturaCartaoForm
    template_name_suffix = '_criar_form'

    def get_success_url(self):
        # Verifica a origem da solicitação na sessão
        origin = self.request.GET.get('fatura_create_origin')
        if origin == 'listagemfaturas':
            return reverse_lazy('listagemfaturas')
        else:
            cartao_pk = self.object.cartao.pk
            return reverse('cartao_visualizar', kwargs={'pk': cartao_pk})


@method_decorator(login_required, name='dispatch')
class FaturaDeleteView(DeleteView):
    model = FaturaCartao
    success_url = reverse_lazy('listagemfaturas')


@method_decorator(login_required, name='dispatch')
class FaturaUpdateView(UpdateView):
    model = FaturaCartao
    form_class = FaturaCartaoForm
    template_name_suffix = "_criar_form"
    success_url = reverse_lazy('listagemfaturas')


@method_decorator(login_required, name='dispatch')
class FaturaListView(ListView):
    template_name = 'convenios/listagem_faturas.html'
    model = FaturaCartao
    context_object_name = 'list_objs'
    paginate_by = 5
    def get_queryset(self):
        queryset = super().get_queryset()
        nome_pesquisado = self.request.GET.get('obj')
        if nome_pesquisado:
            queryset = queryset.filter(cartao__titular__nomecompleto__icontains=nome_pesquisado).order_by('cartao')
        else:
            queryset = queryset.all().order_by('cartao')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@method_decorator(login_required, name='dispatch')
class ContratoPlanoOdontologicoCreateView(CreateView):
    form_class = ContratoPlanoOdontologicoForm
    template_name = 'convenios/contratacaoplanoodontologico_criar_form.html'
    success_url = reverse_lazy('listagemcontratoodontologica')



@method_decorator(login_required, name='dispatch')
class ContratoPlanoOdontologicoDetailView(DetailView):
    model = ContratoPlanoOdontologico
    template_name = 'convenios/contratoplanoOdontologico_detalhes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contrato_pk = self.kwargs['pk']
        contrato = ContratoPlanoOdontologico.objects.get(pk=contrato_pk)
        context['contrato'] = contrato
        return context



@method_decorator(login_required, name='dispatch')
class ContratoPlanoOdontologicoUpdateView(UpdateView):
    model = ContratoPlanoOdontologico
    form_class = ContratoPlanoOdontologicoForm
    template_name = 'convenios/contratacaoplanoodontologico_criar_form.html'
    success_url = reverse_lazy('listagemcontratoodontologica')


@method_decorator(login_required, name='dispatch')
class ContratoPlanoOdontologicoDeleteView(DeleteView):
    model = ContratoPlanoOdontologico
    success_url = reverse_lazy('listagemcontratoodontologica')


@method_decorator(login_required, name='dispatch')
class ContratoOdontologicaListView(ListView):
    template_name = 'convenios/listagem_contratacaoodontologica.html'
    model = ContratoPlanoOdontologico
    context_object_name = 'list_objs'
    paginate_by = 5

    def get_queryset(self):
        queryset = super().get_queryset()
        nome_pesquisado = self.request.GET.get('obj')
        if nome_pesquisado:
            queryset = queryset.filter(contratante__nomecompleto__icontains=nome_pesquisado).order_by('contratante')
        else:
            queryset = queryset.order_by('contratante')
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class VerificarAssociacaoView(View):
    @csrf_exempt
    def get(self, request):
        contratante_id = self.request.GET.get('contratante_id')
        plano_odontologico_id = self.request.GET.get('plano_odontologico_id')
        quantidade_dependentes = self.request.GET.get('quantidade_dependentes')

        print(contratante_id, plano_odontologico_id, quantidade_dependentes)

        contratante = Associado.objects.get(pk=contratante_id)
        taxa_administrativa = TaxasAdministrativa.objects.get(grupos=contratante.associacao)
        percentual_taxa = taxa_administrativa.percentual

        valor_unitario = PlanoOdontologico.objects.get(id=plano_odontologico_id).valorUnitario


        # instance.valor = (valorPlano / (Decimal(100.0) - taxa.percentual)) * Decimal(100.0)
        valor_total = round(((valor_unitario + quantidade_dependentes + 1) / (100 - percentual_taxa)) * 100, 2)

        print(valor_unitario, percentual_taxa, valor_total)

        return JsonResponse({'valor_total': valor_total})


class VerificarDependentesView(View):
    @csrf_exempt
    def get(self, request):
        contratante_id = self.request.GET.get('contratante_id')
        contratante_id = request.GET.get('contratante_id')
        try:
            contratante = Associado.objects.get(pk=contratante_id)
            dependentes = contratante.dependentes.all()
            dependentes_data = [{'id': dep.pk, 'nomecompleto': dep.nomecompleto} for dep in dependentes]
            return JsonResponse({'has_dependents': dependentes.exists(), 'dependentes': dependentes_data})
        except Associado.DoesNotExist:
            return JsonResponse({'has_dependents': False, 'dependentes': []})


@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeCreateView(CreateView):
    model = ContratoPlanoSaude
    form_class = ContratoPlanoSaudeForm
    template_name = 'convenios/planosaude/contrato_plano_saude_add.html'
    success_url = reverse_lazy('listar_contratos_plano_saude')

    def form_valid(self, form):
        contrato = form.save(commit=False)
        taxa_administrativa = TaxasAdministrativa.objects.get(grupos=contrato.contratante.associacao)
        valor_faixa = ValoresPorFaixa.objects.get(pk=contrato.faixaEtaria.id)
        percentual_taxa = taxa_administrativa.percentual
        contrato.valorTotal = round((valor_faixa / (100 - percentual_taxa)) * 100, 2)
        contrato.save()
        return super().form_valid(form)

@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeListView(ListView):
    template_name = 'convenios/planosaude/listagem_contratos_plano_saude.html'
    model = ContratoPlanoSaude
    context_object_name = 'list_objs'
    paginate_by = 5

    def get_queryset(self):
        queryset = super().get_queryset()
        nome_pesquisado = self.request.GET.get('obj')
        if nome_pesquisado:
            queryset = queryset.filter(contratante__nomecompleto__icontains=nome_pesquisado).order_by('contratante')
        else:
            queryset = queryset.order_by('contratante')
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeUpdateView(UpdateView):
    model = ContratoPlanoSaude
    form_class = ContratoPlanoSaudeForm
    template_name = 'convenios/planosaude/contrato_plano_saude_add.html'
    success_url = reverse_lazy('listar_contratos_plano_saude')

@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeDeleteView(DeleteView):
    model = ContratoPlanoSaude
    success_url = reverse_lazy('listar_contratos_plano_saude')

@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeDetailView(DetailView):
    model = ContratoPlanoSaude
    template_name = 'convenios/planosaude/button_model_contrato_plano_saude_detail.html'
    context_object_name = 'contrato'

@method_decorator(login_required, name='dispatch')
class ConsultaValorFaixaEtaria(View):
    @csrf_exempt
    def get(self, request):
        planoSaude_id = self.request.GET.get('planoSaude_id')
        contratante_id = self.request.GET.get('contratante_id')
        atendimentoDomiciliar = self.request.GET.get('atendimentoDomiciliar')

        idade = self.calcular_idade(Associado.objects.get(pk=contratante_id).dataNascimento)

        faixa = ValoresPorFaixa.objects.filter(planoSaude_id=planoSaude_id, idadeMin__lte=idade, idadeMax__gte=idade).first()
        valorAtendimentoDomiciliar = PlanoSaude.objects.get(pk=planoSaude_id).valorAtendimentoDomiciliar
        valorAtendimentoTelefonico = PlanoSaude.objects.get(pk=planoSaude_id).valorAtendimentoTelefonico

        taxa = TaxasAdministrativa.objects.get(grupos = Associado.objects.get(pk=contratante_id).associacao)

        if atendimentoDomiciliar == 'True':
            valor = round(((faixa.valor+valorAtendimentoTelefonico) / (100-taxa.percentual)) * 100,2) + round((valorAtendimentoDomiciliar/(100-taxa.percentual)*100),2)
        else:
            valor = round(((faixa.valor+valorAtendimentoTelefonico)  / (100-taxa.percentual)) * 100,2)

        return JsonResponse({'faixa_id': faixa.id, 'valor': valor,
                             'idadeMin': faixa.idadeMin, 'idadeMax': faixa.idadeMax
                             })

    def calcular_idade(self, dataNascimento):
        data_atual = datetime.date.today()
        idade = data_atual.year - dataNascimento.year - ((data_atual.month, data_atual.day) < (dataNascimento.month, dataNascimento.day))
        return idade
@method_decorator(login_required, name='dispatch')
class ContratoPlanoSaudeDependenteCreateView(CreateView):
    model = ContratoPlanoSaudeDependente
    form_class = ContratoPlanoSaudeDependenteForm
    template_name = 'convenios/planosaude/contrato_plano_dependente_saude_add.html'
    success_url = reverse_lazy('listar_contratos_plano_saude')

    def form_valid(self, form):

        dependente_contrato = form.save(commit=False)

        taxa_administrativa = TaxasAdministrativa.objects.get(grupos=dependente_contrato.contrato.contratante.associacao)

        valor_faixa = ValoresPorFaixa.objects.get(pk=dependente_contrato.faixa.id)

        print(valor_faixa.valor         , type(valor_faixa.valor))

        percentual_taxa = taxa_administrativa.percentual

        dependente_contrato.valorTotal = round((valor_faixa.valor / (Decimal(100) - percentual_taxa)) * 100, 2)

        dependente_contrato.save()

        titular_contrato = get_object_or_404(ContratoPlanoSaude, )
        titular_contrato.valorTotal = titular_contrato.valorTotal + dependente_contrato.valorTotal
        titular_contrato.save()

        return super().form_valid(form)

@method_decorator(login_required, name='dispatch')
class ConsultaValorFaixaEtariaDependente(View):
    @csrf_exempt
    def get(self, request):
        planoSaude_id = self.request.GET.get('planoSaude_id')
        dependente_id = self.request.GET.get('dependente_id')
        atendimentoDomiciliar = self.request.GET.get('atendimentoDomiciliar')

        print (planoSaude_id, dependente_id, atendimentoDomiciliar)
        idade = self.calcular_idade(Dependente.objects.get(pk=dependente_id).dataNascimento)

        faixa = ValoresPorFaixa.objects.filter(planoSaude_id=planoSaude_id, idadeMin__lte=idade, idadeMax__gte=idade).first()
        valorAtendimentoDomiciliar = PlanoSaude.objects.get(pk=planoSaude_id).valorAtendimentoDomiciliar
        valorAtendimentoTelefonico = PlanoSaude.objects.get(pk=planoSaude_id).valorAtendimentoTelefonico

        taxa = TaxasAdministrativa.objects.get(grupos = Dependente.objects.get(pk=dependente_id).titular.associacao)

        if atendimentoDomiciliar == 'True':
            valor = round(((faixa.valor+valorAtendimentoTelefonico) / (100-taxa.percentual)) * 100,2) + round((valorAtendimentoDomiciliar/(100-taxa.percentual)*100),2)
        else:
            valor = round(((faixa.valor+valorAtendimentoTelefonico)  / (100-taxa.percentual)) * 100,2)

        return JsonResponse({'faixa_id': faixa.id, 'valor': valor,
                             'idadeMin': faixa.idadeMin, 'idadeMax': faixa.idadeMax
                             })

    def calcular_idade(self, dataNascimento):
        data_atual = datetime.date.today()
        idade = data_atual.year - dataNascimento.year - ((data_atual.month, data_atual.day) < (dataNascimento.month, dataNascimento.day))
        return idade

def exportar(request):
    empresa_selecionada = request.GET.get('inputGroupSelectEmpresa')
    data_inicial = request.GET.get('start_date')
    data_final = request.GET.get('end_date')
    tipoArquivo_selecionado = request.GET.get('inputGroupSelectTipoArquivo')

    print(f"{empresa_selecionada}, {data_inicial}, {data_final}")

    if data_inicial and data_final and empresa_selecionada and tipoArquivo_selecionado:
        # Filtra as faturas pelo intervalo de datas e pelo nome da empresa
        if empresa_selecionada == '5':
            faturas = FaturaCartao.objects.filter(
                competencia__gte=data_inicial, competencia__lte=data_final
            ).order_by('competencia')

            nome_arq = 'faturas_{}_{}'.format(data_inicial, data_final)
        else:
            faturas = FaturaCartao.objects.filter(
                cartao__titular__empresa_id=empresa_selecionada,
                competencia__gte=data_inicial, competencia__lte=data_final
            ).order_by('competencia')
            nome_arq = 'fatura_{}_{}'.format(data_inicial, data_final)

        return exporttofile(faturas=faturas,
                            nome_arq=nome_arq,
                            tipoArquivo_selecionado=tipoArquivo_selecionado)

    return reverse_lazy('listagemfaturas')


def exporttofile(faturas, nome_arq, tipoArquivo_selecionado):
    global response
    filename = f"{nome_arq}"

    if tipoArquivo_selecionado == '1':

        quant_faturas = faturas.count()
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        y = 750  # Posição y inicial
        for fatura in faturas:
            if y <= 50:
                pdf.showPage()
                y = 750
            # Escreve o título da fatura
            pdf.drawString(100, y, "Titular do Cartão: {}".format(fatura.cartao.titular.nomecompleto))
            y -= 20  # Move para a próxima linha
            pdf.drawString(100, y, "Empresa: {}".format(fatura.cartao.titular.empresa.nome))
            y -= 20  # Move para a próxima linha
            # Escreve as informações da fatura
            pdf.drawString(100, y, "Competência: {}".format(fatura.competencia))
            y -= 20
            pdf.drawString(100, y, "Valor: {}".format(fatura.valorComTaxa))
            y -= 40  # Move duas linhas para baixo

        pdf.save()
        # Define o nome do arquivo PDF
        filename = filename + ".pdf"
        # Envia o PDF para o navegador como um arquivo de download
        response = HttpResponse(content_type='application/pdf')

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        buffer.seek(0)

        response.write(buffer.read())

        buffer.close()
    elif tipoArquivo_selecionado == '2':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nome_arq
        # Escreve os dados da lista de faturas na planilha
        ws['A1'] = "Titular"
        ws['B1'] = "Empresa"
        ws['C1'] = "Competêcia"
        ws['D1'] = "Valor"


        for i, fatura in enumerate(faturas):
            ws.cell(row=i + 2, column=1, value=fatura.cartao.titular.nomecompleto)
            ws.cell(row=i + 2, column=2, value=fatura.cartao.titular.empresa.nome)
            ws.cell(row=i + 2, column=3, value=fatura.competencia)
            ws.cell(row=i + 2, column=5, value=fatura.valorComTaxa)

        # Cria uma resposta HTTP com o arquivo XLSX
        response = HttpResponse(content_type='application/xlsx')
        response['Content-Disposition'] = f'attachment; filename={nome_arq}.xlsx'
        wb.save(response)


    elif tipoArquivo_selecionado == '3':
        faturas_txt = ''
        for fatura in faturas:
            matricula = str(fatura.cartao.titular.matricula).zfill(6)
            valorComTaxa = str(fatura.valorComTaxa)[:4].replace('.', '').zfill(12)
            faturas_txt += '{} {}\n'.format(matricula, valorComTaxa)

        response = HttpResponse(faturas_txt, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{nome_arq}.txt'

    return response
