from io import BytesIO
import openpyxl
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import TemplateView
from reportlab.pdfgen import canvas
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from blackpearl.convenios.models import FaturaCartao

@method_decorator(login_required, name='dispatch')
class HomeTemplateView(TemplateView):
    template_name = 'convenios/home.html'

    def get_context_data(self, **kwargs):
        context = super(HomeTemplateView, self).get_context_data(**kwargs)
        return context


def exportar(request):
    empresa_selecionada = request.GET.get('inputGroupSelectEmpresa')
    data_inicial = request.GET.get('start_date')
    data_final = request.GET.get('end_date')
    tipoArquivo_selecionado = request.GET.get('inputGroupSelectTipoArquivo')

    print(f"{empresa_selecionada}, {data_inicial}, {data_final}")

    if data_inicial and data_final and empresa_selecionada and tipoArquivo_selecionado:
        # Filtra as faturas pelo intervalo de datas e pelo nome da empresa
        if empresa_selecionada == '5':
            faturas = FaturaCartao.objects.filter(
                competencia__gte=data_inicial, competencia__lte=data_final
            ).order_by('competencia')

            nome_arq = 'faturas_{}_{}'.format(data_inicial, data_final)
        else:
            faturas = FaturaCartao.objects.filter(
                cartao__titular__empresa_id=empresa_selecionada,
                competencia__gte=data_inicial, competencia__lte=data_final
            ).order_by('competencia')
            nome_arq = 'fatura_{}_{}'.format(data_inicial, data_final)

        return exporttofile(faturas=faturas,
                            nome_arq=nome_arq,
                            tipoArquivo_selecionado=tipoArquivo_selecionado)

    return reverse_lazy('listagemfaturas')


def exporttofile(faturas, nome_arq, tipoArquivo_selecionado):
    global response
    filename = f"{nome_arq}"

    if tipoArquivo_selecionado == '1':

        quant_faturas = faturas.count()
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        y = 750  # Posição y inicial
        for fatura in faturas:
            if y <= 50:
                pdf.showPage()
                y = 750
            # Escreve o título da fatura
            pdf.drawString(100, y, "Titular do Cartão: {}".format(fatura.cartao.titular.nomecompleto))
            y -= 20  # Move para a próxima linha
            pdf.drawString(100, y, "Empresa: {}".format(fatura.cartao.titular.empresa.nome))
            y -= 20  # Move para a próxima linha
            # Escreve as informações da fatura
            pdf.drawString(100, y, "Competência: {}".format(fatura.competencia))
            y -= 20
            pdf.drawString(100, y, "Valor: {}".format(fatura.valorComTaxa))
            y -= 40  # Move duas linhas para baixo

        pdf.save()
        # Define o nome do arquivo PDF
        filename = filename + ".pdf"
        # Envia o PDF para o navegador como um arquivo de download
        response = HttpResponse(content_type='application/pdf')

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        buffer.seek(0)

        response.write(buffer.read())

        buffer.close()
    elif tipoArquivo_selecionado == '2':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nome_arq
        # Escreve os dados da lista de faturas na planilha
        ws['A1'] = "Titular"
        ws['B1'] = "Empresa"
        ws['C1'] = "Competêcia"
        ws['D1'] = "Valor"


        for i, fatura in enumerate(faturas):
            ws.cell(row=i + 2, column=1, value=fatura.cartao.titular.nomecompleto)
            ws.cell(row=i + 2, column=2, value=fatura.cartao.titular.empresa.nome)
            ws.cell(row=i + 2, column=3, value=fatura.competencia)
            ws.cell(row=i + 2, column=5, value=fatura.valorComTaxa)

        # Cria uma resposta HTTP com o arquivo XLSX
        response = HttpResponse(content_type='application/xlsx')
        response['Content-Disposition'] = f'attachment; filename={nome_arq}.xlsx'
        wb.save(response)


    elif tipoArquivo_selecionado == '3':
        faturas_txt = ''
        for fatura in faturas:
            matricula = str(fatura.cartao.titular.matricula).zfill(6)
            valorComTaxa = str(fatura.valorComTaxa)[:4].replace('.', '').zfill(12)
            faturas_txt += '{} {}\n'.format(matricula, valorComTaxa)

        response = HttpResponse(faturas_txt, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{nome_arq}.txt'

    return response
